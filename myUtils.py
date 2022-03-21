#!/usr/bin/env python
# coding=utf-8
import datetime
import os
import re
import shutil

import openpyxl as oxl
import requests
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QTableWidgetItem
from bs4 import BeautifulSoup

# 全局变量区域
borderNumDic = {-1: None, 0: "thin"}

# 访问URL,type表示请求类型，0为GET，1为POST，2为PUT
# 返回值类型如下：
# {
# "url":传入URL,
# "resultStr":访问结果字符串,
# "checkFlag":标志是否访问成功的布尔类型变量,
# "title":访问成功时的页面标题,
# "pageContent":访问成功时的页面源码，
# "status":访问的响应码，
# "requestSeconed":访问耗时，单位为秒
# }
from openpyxl.styles import Border, Side, Font, PatternFill


def requestsUrl(url, cookie={}, header={}, data={}, files=None, type=0, reqTimeout=10, readTimeout=10,
                allow_redirects=False, session=None, proxy=""):
    proxies = None
    if proxy != "":
        proxies = {
            'http': 'http://' + proxy,
            'https': 'https://' + proxy
        }
    else:
        pass
    if session is None:
        session = requests.session()
    else:
        pass
    resDic = {}
    url = url.strip()
    url = url.strip()

    resultStr = ""
    checkedFlag = False
    status = ""
    title = ""
    reContent = ""
    totalSeconed = 0
    timeout = (reqTimeout, readTimeout)
    header = header if header != {} else {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; WOW64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/75.0.3770.100 Safari/537.36"
    }
    try:
        if type == 0:
            response = session.get(url, headers=header, verify=False, cookies=cookie, timeout=timeout,
                                   allow_redirects=allow_redirects, proxies=proxies)
        elif type == 1:
            response = session.post(url, headers=header, verify=False, cookies=cookie, data=data, files=files,
                                    timeout=timeout, allow_redirects=allow_redirects, proxies=proxies)
        elif type == 2:
            response = session.put(url, headers=header, verify=False, cookies=cookie, data=data, files=files,
                                   timeout=timeout, allow_redirects=allow_redirects, proxies=proxies)
        else:
            pass
        status = response.status_code
        totalSeconed = response.elapsed.total_seconds()
        if str(status)[0] == "2" or str(status)[0] == "3":
            # 获得页面编码
            pageEncoding = response.apparent_encoding
            # 设置页面编码
            response.encoding = pageEncoding
            # 获得页面内容
            reContent = response.text
            soup = BeautifulSoup(reContent, "lxml")
            title = "成功访问，但无法获得标题" if not soup.title else soup.title.string
            resultStr = "验证成功，标题为：{0}".format(title)
            checkedFlag = True
        else:
            resultStr = "验证失败，状态码为{0}".format(status)
            # 获得页面编码
            pageEncoding = response.apparent_encoding
            # 设置页面编码
            response.encoding = pageEncoding
            # 获得页面内容
            reContent = response.text
            checkedFlag = False
    except Exception as e:
        resultStr = str(e)
        checkedFlag = False

    # 构建返回结果
    resDic["url"] = url
    resDic["resultStr"] = resultStr
    resDic["checkFlag"] = checkedFlag
    resDic["title"] = title
    resDic["status"] = status
    resDic["pageContent"] = reContent
    resDic["requestSeconed"] = totalSeconed
    return resDic


# 获得精确到秒的当前时间
def getNowSeconed():
    formatStr = "%Y-%m-%d %H:%M:%S"
    nowDate = datetime.datetime.now()
    nowDateStr = nowDate.strftime(formatStr)
    return nowDateStr


# 将传入的字符串修改为符合windows文件名规范的字符串
def updateFileNameStr(oriStr):
    resultStr = oriStr
    # 替换换行符
    resultStr = resultStr.replace("\r\n", "\n").replace("\n", "")
    # 将违法字符替换为下划线
    notAllowCharList = ["\\", "/", ":", "*", "?", "\"", "<", ">", "|"]
    for nowNotAllowChar in notAllowCharList:
        resultStr = resultStr.replace(nowNotAllowChar, "_")
    return resultStr


# 在本目录创建一个文件夹,ifDelete 表示如果已经存在该文件夹是否会删除原有文件夹并重新创建
# 返回一个布尔变量，表示是否创建了文件夹
def createFolderHere(folderName, ifDelete=True):
    basePath = os.getcwd()
    ifCreate = True
    if os.path.exists(folderName):
        if ifDelete:
            shutil.rmtree(basePath + "\\" + folderName)
            os.makedirs(basePath + "\\" + folderName)
        else:
            ifCreate = False
    else:
        # 不存在文件夹，创建文件夹
        os.makedirs(basePath + "\\" + folderName)

    return ifCreate


def checkIfUrl(urlStr):
    checkFlag = True
    urlRegex = re.compile(
        r'^(?:http|ftp)s?://'  # http:// or https://
        r'(?:(?:[A-Z0-9](?:[A-Z0-9-]{0,61}[A-Z0-9])?\.)+(?:[A-Z]{2,6}\.?|[A-Z0-9-]{2,}\.?)|'  # domain...
        r'localhost|'  # localhost...
        r'\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3})'  # ...or ip
        r'(?::\d+)?'  # optional port
        r'(?:/?|[/?]\S+)$', re.IGNORECASE)
    checkFlag = urlRegex.match(urlStr)
    return checkFlag


# 传入一个QTableWidget对象，清空这个表格
def clearTalbe(tableObject):
    while tableObject.rowCount() != 0:
        tableObject.removeRow(tableObject.rowCount() - 1)


# 创建一个QTableWidgetItem对象，传入文本参数text，以及verAlign和horAlign两个整形参数控制文本位置
# verAlign的值含义如下：-1 靠左、0 居中、1 靠右
# horAlign的值含义如下：-1 靠下、0 居中、1 靠上
def createTableItem(text, verAlign=0, horAlign=0):
    tmpItem = QTableWidgetItem(str(text))
    verAlignDic = {-1: Qt.AlignLeft, 0: Qt.AlignVCenter, 1: Qt.AlignRight}
    horAlignDic = {-1: Qt.AlignBottom, 0: Qt.AlignHCenter, 1: Qt.AlignTop}
    tmpItem.setTextAlignment(verAlignDic[verAlign] | horAlignDic[horAlign])
    return tmpItem


# 删除指定路径的文件,传入一个绝对路径,返回一个布尔变量以及一个字符串变量，
# 布尔变量为True表示是否删除成功,若为False则字符串变量中写入错误信息
def deleteFile(filePath):
    deleteFlag = True
    reStr = ""
    if os.path.exists(filePath):
        try:
            os.remove(filePath)
        except Exception as ex:
            reStr = "删除失败，失败信息为：{0}".format(ex)
            deleteFlag = False
    else:
        reStr = "未找到指定路径的文件"
        deleteFlag = False
    return deleteFlag, reStr


# 获得excell的常用样式
def getExcellStyleDic():
    styleDic = {}

    # 单线边框
    thinBorder = Border(left=Side(border_style='thin', color='000000'),
                        right=Side(border_style='thin', color='000000'),
                        top=Side(border_style='thin', color='000000'),
                        bottom=Side(border_style='thin', color='000000'))

    # 文字居中
    alignStyle = oxl.styles.Alignment(horizontal='center', vertical='center')
    leftStyle = oxl.styles.Alignment(horizontal='left', vertical='center')
    rightStyle = oxl.styles.Alignment(horizontal='right', vertical='center')

    # 加粗字体
    boldFont = Font(bold=True)
    hyperLinkFont = Font(color='0000FF')
    underLineFont = Font(underline='single')

    styleDic["thin"] = thinBorder
    styleDic["align"] = alignStyle
    styleDic["bold"] = boldFont
    styleDic["left"] = leftStyle
    styleDic["right"] = rightStyle
    styleDic["link"] = hyperLinkFont
    styleDic["underLine"] = underLineFont
    return styleDic


# 写入一个标准的excell表头（居中，单线框，加粗）
def writeExcellHead(ws, headArr):
    # 获得常用样式
    styleDic = getExcellStyleDic()
    # 写入表头
    for index, head in enumerate(headArr):
        ws.cell(row=1, column=index + 1).value = head
        ws.cell(row=1, column=index + 1).border = styleDic["thin"]
        ws.cell(row=1, column=index + 1).alignment = styleDic["align"]
        ws.cell(row=1, column=index + 1).font = styleDic["bold"]
    return ws


# 写入一个内容单元格
# borderNum表示该单元格的边框对象，其值可查询全局变量styleDic
# ifAlign是一个boolean对象，True表示居中
# hyperLink表示该单元格指向的链接，默认为None，表示不指向任何链接
# fgColor表示该单元格的背景颜色，为一个RGB16进制字符串，默认为“FFFFFF”（白色）
# otherAlign表示当ifAlign为False时指定的其他对齐方式，是一个数字型变量，默认为None，当其为0时表示左对齐，1为右对齐
def writeExcellCell(ws, row, column, value, borderNum, ifAlign, hyperLink=None, fgColor="FFFFFF", otherAlign=None):
    # 获得常用样式
    styleDic = getExcellStyleDic()
    # 获得指定单元格
    aimCell = ws.cell(row=row, column=column)
    # 设置值
    aimCell.value = value
    # 设置边框
    styleObjKey = borderNumDic[borderNum]
    if not styleObjKey:
        pass;
    else:
        styleObj = styleDic[styleObjKey]
        aimCell.border = styleObj
    # 设置居中
    if ifAlign:
        aimCell.alignment = styleDic["align"]
    elif otherAlign is not None:
        otherAlign = int(otherAlign)
        if otherAlign == 0:
            aimCell.alignment = styleDic["left"]
        else:
            aimCell.alignment = styleDic["right"]
    else:
        pass

    # 设置超链接
    if hyperLink:
        # 写入超链接
        aimCell.hyperlink = hyperLink
        # 设置当前单元格字体颜色为深蓝色，并添加下划线
        aimCell.font = styleDic["link"]
    else:
        pass

    # 设置填充颜色
    fill = PatternFill("solid", fgColor=fgColor)
    aimCell.fill = fill

    return ws


# 写入一个空格单元格，防止上一列文本超出
def writeExcellSpaceCell(ws, row, column):
    # 设置值
    ws.cell(row=row, column=column).value = " "

    return ws


# 设置excell的列宽
def setExcellColWidth(ws, colWidthArr):
    for colWidindex in range(len(colWidthArr)):
        ws.column_dimensions[chr(ord("A") + colWidindex)].width = colWidthArr[colWidindex]

    return ws


# 保存excell文件
def saveExcell(wb, saveName):
    savePath = ""
    # 处理传入的文件名
    saveName = saveName.split(".")[0] + ".xlsx"
    savePath = "{0}\\{1}".format(os.getcwd(), saveName)

    # 检测当前目录下是否有该文件，如果有则清除以前保存文件
    if os.path.exists(savePath):
        deleteFile(savePath)
    wb.save(savePath)
    return True


def writeToFile(content, fileName):
    filePath = os.path.join(os.getcwd(), fileName)
    with open(fileName, "w+", encoding="utf-8") as fr:
        fr.write(content)
    return filePath


def changeJsToString(jsStr):
    # 将传入的JS函数字符串转换为字符串常量格式
    resultStr = ""
    # 处理注释和空行
    tmpList = jsStr.split("\n")
    tmpSolvedList = []
    for nowLine in tmpList:
        tmpLine = nowLine
        while len(tmpLine) != 0 and tmpLine[0] == " ":
            tmpLine = tmpLine[1:]
        if len(tmpLine) == 0 or tmpLine[:2] == "//":
            continue
        else:
            tmpSolvedList.append(nowLine)
    jsStr = "\n".join(tmpSolvedList)

    resultStr = jsStr.replace("{\n", "{").replace("}\n", "}").replace(";\n", ";").replace("\n", "\\\\n").replace("\\n",
                                                                                                                 "\\\\n").replace(
        "\"", "\\\"")
    return resultStr
